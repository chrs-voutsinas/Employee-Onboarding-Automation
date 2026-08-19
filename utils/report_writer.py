from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


def write_report(results, file_path):
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = "Employee Results"

    # Business-friendly report headers
    headers = [
        "First Name",
        "Middle Name",
        "Last Name",
        "Employee ID",
        "Status",
        "Error Message"
    ]

    worksheet.append(headers)

    # Write employee results
    for result in results:
        worksheet.append([
            result.get("first_name", ""),
            result.get("middle_name", ""),
            result.get("last_name", ""),
            result.get("employee_id", ""),
            result.get("status", ""),
            result.get("error_message", "")
        ])

    # Header formatting
    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78"
    )

    header_font = Font(
        color="FFFFFF",
        bold=True
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    # Keep headers visible while scrolling
    worksheet.freeze_panes = "A2"

    # Status formatting
    success_fill = PatternFill(
        fill_type="solid",
        fgColor="C6EFCE"
    )

    failure_fill = PatternFill(
        fill_type="solid",
        fgColor="FFC7CE"
    )

    for row in range(2, worksheet.max_row + 1):
        status_cell = worksheet.cell(
            row=row,
            column=5
        )

        if status_cell.value == "Success":
            status_cell.fill = success_fill

        elif status_cell.value == "Failed":
            status_cell.fill = failure_fill

    # Create Excel table with filters
    if results:
        table_reference = (
            f"A1:"
            f"{get_column_letter(worksheet.max_column)}"
            f"{worksheet.max_row}"
        )

        table = Table(
            displayName="EmployeeResults",
            ref=table_reference
        )

        table_style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )

        table.tableStyleInfo = table_style
        worksheet.add_table(table)

    # Business-friendly column widths
    column_widths = {
        "A": 18,
        "B": 18,
        "C": 22,
        "D": 16,
        "E": 14,
        "F": 50
    }

    for column, width in column_widths.items():
        worksheet.column_dimensions[column].width = width

    # Improve readability
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True
            )

    # Ensure output directory exists
    file_path.parent.mkdir(
        parents=True,
        exist_ok=True
    )

    workbook.save(file_path)