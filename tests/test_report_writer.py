from openpyxl import load_workbook

from utils.report_writer import write_report


def test_write_report_creates_excel_file(tmp_path):
    results = [
        {
            "first_name": "John",
            "middle_name": "",
            "last_name": "Smith",
            "employee_id": "12345",
            "status": "Success",
            "error_message": "",
            "screenshot_path": "screenshots/test.png"
        }
    ]

    report_path = tmp_path / "employee_results.xlsx"

    write_report(
        results,
        report_path
    )

    assert report_path.exists()


def test_write_report_contains_business_columns(tmp_path):
    results = [
        {
            "first_name": "John",
            "middle_name": "",
            "last_name": "Smith",
            "employee_id": "12345",
            "status": "Success",
            "error_message": "",
            "screenshot_path": "screenshots/test.png"
        }
    ]

    report_path = tmp_path / "employee_results.xlsx"

    write_report(
        results,
        report_path
    )

    workbook = load_workbook(report_path)
    worksheet = workbook["Employee Results"]

    headers = [
        cell.value
        for cell in worksheet[1]
    ]

    assert headers == [
        "First Name",
        "Middle Name",
        "Last Name",
        "Employee ID",
        "Status",
        "Error Message"
    ]


def test_write_report_excludes_screenshot_path(tmp_path):
    results = [
        {
            "first_name": "Alex",
            "middle_name": "",
            "last_name": "Brown",
            "employee_id": "",
            "status": "Failed",
            "error_message": "Employee creation failed",
            "screenshot_path": "screenshots/failure.png"
        }
    ]

    report_path = tmp_path / "employee_results.xlsx"

    write_report(
        results,
        report_path
    )

    workbook = load_workbook(report_path)
    worksheet = workbook["Employee Results"]

    all_values = [
        cell.value
        for row in worksheet.iter_rows()
        for cell in row
    ]

    assert "Screenshot Path" not in all_values
    assert "screenshots/failure.png" not in all_values


def test_write_report_contains_employee_results(tmp_path):
    results = [
        {
            "first_name": "John",
            "middle_name": "",
            "last_name": "Smith",
            "employee_id": "12345",
            "status": "Success",
            "error_message": ""
        },
        {
            "first_name": "Alex",
            "middle_name": "",
            "last_name": "Brown",
            "employee_id": "",
            "status": "Failed",
            "error_message": "Employee creation failed"
        }
    ]

    report_path = tmp_path / "employee_results.xlsx"

    write_report(
        results,
        report_path
    )

    workbook = load_workbook(report_path)
    worksheet = workbook["Employee Results"]

    assert worksheet["A2"].value == "John"
    assert worksheet["C2"].value == "Smith"
    assert worksheet["D2"].value == "12345"
    assert worksheet["E2"].value == "Success"

    assert worksheet["A3"].value == "Alex"
    assert worksheet["C3"].value == "Brown"
    assert worksheet["E3"].value == "Failed"
    assert worksheet["F3"].value == "Employee creation failed"